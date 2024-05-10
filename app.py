from flask import Flask, request, jsonify
import numpy as np
from openai import OpenAI
import pytesseract as pyt
import cv2
import openpyxl
import os

app = Flask(__name__)



# Set Tesseract path
#pyt.pytesseract.tesseract_cmd = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

UPLOAD_FOLDER = '/app/imgs'
ALLOWED_EXTENSIIONS=set(['png','jpg','jpeg','gif'])

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/extract', methods=['POST'])
#@app.route('/find')
# def home():
#     return"hello world"
def extract_info():

    #return jsonify('entered')
    if 'image' not in request.files:
        return jsonify({'error':'image not provided'}),400
    # Ensure the temp directory exists
    # if not os.path.exists("temp"):
    #     os.makedirs("temp")

    # Receive image file
    image_file = request.files['image']

    # Save the uploaded file to a temporary location
    # temp_file_path = os.path.join("temp", image_file.filename)
    # image_file.save(temp_file_path)

    image_file.save(os.path.join(app.config['UPLOAD_FOLDER'],image_file.filename))


    # Read image
    image = cv2.imread(os.path.join(app.config['UPLOAD_FOLDER'],image_file.filename))

    #return jsonify('image readed')

    # Extract text from image
    text = pyt.image_to_string(image)

    #return jsonify(text)
    # Define context
    context = "I am a finder. I can extract information from government proofs for my personal use. I'll print the details of the proof in the given format.To choose the correct provider, I'll check if the number is a 12-digit number then it is Aadhar, if it is  10-digit alphanumeric text then it is PAN and if it is a  10 digit alphanumeric text and contains election commision of india in it then it a Voter ID. Based on the provider the output format differs Output format for Aadhar: 1$<number>$<Name>$<DOB>$<gender> ,Output format for PAN: 2$<PAN account number>$<Name>$<Father name>$<DOB> ,Output format for VoterId: 3$<number>$<Name>$<Fathers name>$<DOB>$<gender>."

    print("context")
    # Generate completion
    
    completion = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": context},
            {"role": "user", "content": text}
        ]
    )


    # Get output text
    output_text = completion.choices[0].message.content

    # Split output into fields
    new_row_data = output_text.split('$')
    
    #return jsonify(new_row_data)

    # Use an environment variable for the Excel file path
    EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH', '/app/data/governmentdetails.xlsx')  

    #"C:\Users\santhosh\Downloads\governmentdetails.xlsx"
    # Load workbook
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)

    # Access or create worksheets
    Aadhar_worksheet = workbook.get_sheet_by_name('Aadhar') if 'Aadhar' in workbook.sheetnames else workbook.create_sheet('Aadhar')
    PAN_worksheet = workbook.get_sheet_by_name('PAN') if 'PAN' in workbook.sheetnames else workbook.create_sheet('PAN')
    VoterId_worksheet = workbook.get_sheet_by_name('VoterId') if 'VoterId' in workbook.sheetnames else workbook.create_sheet('VoterId')

    # Add details based on type
    if new_row_data[0] == '1':
        add_details(Aadhar_worksheet, new_row_data)
    elif new_row_data[0] == '2':
        add_details(PAN_worksheet, new_row_data)
    else:
        add_details(VoterId_worksheet, new_row_data)

    #Save workbook
    workbook.save(EXCEL_FILE_PATH)

    return jsonify({'message': 'Details extracted and saved successfully!' + ', '.join(new_row_data)})

def add_details(worksheet, new_row_data):
    # Find next available row
    next_row = worksheet.max_row + 1

    # Write data to the new row
    for col_num, value in enumerate(new_row_data[1:], start=1): # start=1 to start columns from 1
        worksheet.cell(row=next_row, column=col_num, value=value)

if __name__ == "__main__":
    app.run(host='0.0.0.0', debug=True)
